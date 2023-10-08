#!/usr/bin/python

# ChromeDriver woes: make sure it's up to date https://stackoverflow.com/questions/21166408/how-do-i-confirm-im-using-the-right-chromedriver
# or just rename the .exe at C:\Windows and it maybe just works?

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import datetime
from openpyxl import Workbook
import re
import os

# IsDebugMode = True
IsDebugMode = False

LoadAllPrograms = True
# LoadAllPrograms = False

rootUrl = "https://arts-inspiredlearning.org/programs/"
scrapedResultsFilename = "ArtsInspiredLearningScraped.xlsx"
logFilePath = "C:\src\personalthings\Lyndsey\PythonWebScrapers\LogFile.log"
AllProgramTags = []

if os.path.exists(logFilePath):
  os.remove(logFilePath)

class programDetails:
    routePath = ""
    name = ""
    rawDescription = ""
    formattedDescription = ""
    artists = []
    tags = "" # called program details in the website, Lyndsey wants them as a comma-separated list

    def prettyPrint(self):
        text = f"Name: {self.name}\nRoute Path: {self.routePath}\nRaw Description: {self.rawDescription}\nFormatted Description: {self.formattedDescription}\nArtists: {self.artists}\nTags: {self.tags}"
        print(text)
        return text

def Log(asDebug, message):
    messagePlusNewline = str(message)+"\n"
    if not asDebug:
        print(message)
        f = open(logFilePath, "ab")
        f.write(messagePlusNewline.encode("UTF-8"))
        f.close()
    if asDebug and IsDebugMode:
        print(message)
        f = open(logFilePath, "ab")
        f.write(messagePlusNewline.encode("UTF-8"))
        f.close()

def collectLinksToPrograms(browser):
    containerElement = browser.find_element(By.CLASS_NAME, "bbq-content")
    elements = containerElement.find_elements(By.TAG_NAME, "a")
    Log(True, f"We have {len(elements)} 'a' elements to work with")
    links = []
    for e in elements:
        href = e.get_attribute('href')
        if href is not None:
            links.append(href)
    Log(True, f"We have {len(links)} links to work with")
    return links

def loadMore(browser, previousLinkCount):
    loadMoreDiv = browser.find_element(By.CLASS_NAME, "fl-builder-pagination-load-more")
    loadMoreButton = loadMoreDiv.find_element(By.TAG_NAME, "a")
    if loadMoreButton is not None:
        loadMoreButton.click()
        time.sleep(2)
        linkCount = len(collectLinksToPrograms(browser))
        if linkCount != previousLinkCount:
            Log(True, f"We had {previousLinkCount}, now we have {linkCount}")
            loadMore(browser, linkCount)
    
def getDataForProgram(browser, linkToProgram):
    browser.get(linkToProgram)
    primarySectionElement = browser.find_element(By.ID, "primary")
    programName = primarySectionElement.find_element(By.TAG_NAME, "h1").text
    Log(True, programName)

    descriptionElement = primarySectionElement.find_element(By.CLASS_NAME, "entry-content")
    rawDescription = descriptionElement.text
    Log(True, rawDescription)
    formattedDescription = descriptionElement.get_attribute('innerHTML')
    Log(True, formattedDescription)

    secondarySectionElement = browser.find_element(By.ID, "secondary")
    programArtistsContainerElements = secondarySectionElement.find_elements(By.CLASS_NAME, "sidebar-program-artists")
    artists = []
    if len(programArtistsContainerElements) > 0:
        programArtistsContainerElement = programArtistsContainerElements[0]
        programArtistElements = programArtistsContainerElement.find_elements(By.TAG_NAME, "a")
        for e in programArtistElements:
            artists.append(e.text)
        artists = list( dict.fromkeys(artists) ) # dedupe artists
    Log(True, artists)

    programDetailsContainerElement = secondarySectionElement.find_element(By.CLASS_NAME, "sidebar-program-details")
    items = programDetailsContainerElement.find_elements(By.TAG_NAME, "li")
    Log(True, f"I see {len(items)} items in program details")
    tags = []
    for e in items:
        tagText = e.text
        Log(True, f"Raw tag text: {tagText}")
        badHeaderElements = e.find_elements(By.XPATH, "h4 | h6")
        if len(badHeaderElements) > 0:
            badHeaderElementText = badHeaderElements[0].text
            Log(True, f"badHeaderText: {badHeaderElementText}")
            tagText = tagText.replace(badHeaderElementText, "")
        tags.append(tagText.strip())
    while("" in tags):
        tags.remove("")
    Log(True, tags)
    AllProgramTags.extend(tags)

    program = programDetails()
    program.routePath = linkToProgram.replace(rootUrl, "").rstrip("/")
    program.name = programName
    program.rawDescription = rawDescription
    program.formattedDescription = formattedDescription
    program.artists = artists
    program.tags = ", ".join(tags)
    program.prettyPrint()
    print("\n")
    Log(False, program.prettyPrint())

    return program

def writeHeadersToWorkbook(sheet):
    sheet["A1"] = "Product ID [Non Editable]"
    sheet["B1"] = "Variant ID [Non Editable]"
    sheet["C1"] = "Product Type [Non Editable]"
    sheet["D1"] = "Product Page"
    sheet["E1"] = "Product URL"
    sheet["F1"] = "Title"
    sheet["G1"] = "Description"
    sheet["H1"] = "SKU"
    sheet["I1"] = "Option Name 1"
    sheet["J1"] = "Option Value 1"
    sheet["K1"] = "Option Name 2"
    sheet["L1"] = "Option Value 2"
    sheet["M1"] = "Option Name 3"
    sheet["N1"] = "Option Value 3"
    sheet["O1"] = "Option Name 4"
    sheet["P1"] = "Option Value 4"
    sheet["Q1"] = "Option Name 5"
    sheet["R1"] = "Option Value 5"
    sheet["S1"] = "Option Name 6"
    sheet["T1"] = "Option Value 6"
    sheet["U1"] = "Price"
    sheet["V1"] = "Sale Pricee"
    sheet["W1"] = "On Sale"
    sheet["X1"] = "Stock"
    sheet["Y1"] = "Categories"
    sheet["Z1"] = "Tags"
    sheet["AA1"] = "Weight"
    sheet["AB1"] = "Length"
    sheet["AC1"] = "Width"
    sheet["AD1"] = "Height"
    sheet["AE1"] = "Visible"
    sheet["AF1"] = "Hosted Image URLs"
    # sheet["AG1"] = "The whole thing"
    
def writeProgramToWorkbook(sheet, index, program):
    sheet["A"+str(index)] = ""
    sheet["B"+str(index)] = ""
    sheet["C"+str(index)] = "SERVICE"
    sheet["D"+str(index)] = "arts-programs"
    sheet["E"+str(index)] = program.routePath
    sheet["F"+str(index)] = program.name
    sheet["G"+str(index)] = program.formattedDescription
    sheet["H"+str(index)] = "SQ4218371"
    sheet["I"+str(index)] = ""
    sheet["J"+str(index)] = ""
    sheet["K"+str(index)] = ""
    sheet["L"+str(index)] = ""
    sheet["M"+str(index)] = ""
    sheet["N"+str(index)] = ""
    sheet["O"+str(index)] = ""
    sheet["P"+str(index)] = ""
    sheet["Q"+str(index)] = ""
    sheet["R"+str(index)] = ""
    sheet["S"+str(index)] = ""
    sheet["T"+str(index)] = ""
    sheet["U"+str(index)] = "0"
    sheet["V"+str(index)] = "0"
    sheet["W"+str(index)] = "No"
    sheet["X"+str(index)] = "Unlimited"
    sheet["Y"+str(index)] = program.tags
    sheet["Z"+str(index)] = ""
    sheet["AA"+str(index)] = "0"
    sheet["AB"+str(index)] = "0"
    sheet["AC"+str(index)] = "0"
    sheet["AD"+str(index)] = "0"
    sheet["AE"+str(index)] = "Yes"
    sheet["AF"+str(index)] = ""
    # sheet["AG"+str(index)] = program.prettyPrint()
    index +=1

    return index

Log(False, "Start time: " + datetime.datetime.now())

browser = webdriver.Chrome()
browser.get(rootUrl)

if LoadAllPrograms:
    loadMore(browser, 0)

allProgramLinks = collectLinksToPrograms(browser)
Log(False, allProgramLinks)

workbook = Workbook()
sheet = workbook.active
writeHeadersToWorkbook(sheet)

index = 2
for link in allProgramLinks:
    Log(False, f'{str(index-1)}/{str(len(allProgramLinks))}')
    program = getDataForProgram(browser, link)
    index = writeProgramToWorkbook(sheet, index, program)

workbook.save(filename=scrapedResultsFilename)

browser.quit()

AllProgramTags = list( dict.fromkeys(AllProgramTags) ) # dedupe
print(AllProgramTags)

LOg(False, "End time: " + datetime.datetime.now())